"""Document exports (T7, lot A): relevé, journal and grand livre as PDF / Excel.

Generation is server-side, tenant-scoped and streamed as an attachment. The
tests check the contract (status, media type, attachment headers, valid file
signature), real spreadsheet content (openpyxl), tenant isolation and the
empty-period case. Reading is open to any member; a foreign / non-treasury id
is a 404.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session

from database import get_session
from exports.data import bilan_data, compte_resultat_data
from main import _fastapi_app as app

PASSWORD = "password123"
FROM, TO = "2026-06-01", "2026-06-30"


@pytest.fixture(autouse=True)
def _use_test_session(session: Session):
    app.dependency_overrides[get_session] = lambda: session
    yield
    app.dependency_overrides.clear()


def _client() -> TestClient:
    return TestClient(app)


def _admin_with_association(email: str, name: str) -> tuple[TestClient, str]:
    client = _client()
    assert (
        client.post(
            "/api/auth/register",
            json={"email": email, "password": PASSWORD, "name": "User"},
        ).status_code
        == 201
    )
    assert (
        client.post(
            "/api/auth/login", json={"email": email, "password": PASSWORD}
        ).status_code
        == 200
    )
    resp = client.post(
        "/api/auth/associations", json={"name": name, "email": f"{name}@example.com"}
    )
    assert resp.status_code == 201, resp.text
    return client, resp.json()["id"]


def _categorie_id(client: TestClient, assoc: str, libelle: str) -> str:
    cats = client.get(f"/api/asso/{assoc}/categories").json()
    return next(c["id"] for c in cats if c["libelle"] == libelle)


def _treso_id(client: TestClient, assoc: str, numero: str) -> str:
    comptes = client.get(f"/api/asso/{assoc}/tresorerie").json()
    return next(c["id"] for c in comptes if c["numero"] == numero)


def _compte_id(client: TestClient, assoc: str, numero: str) -> str:
    comptes = client.get(f"/api/asso/{assoc}/comptes", params={"search": numero}).json()
    return next(c["id"] for c in comptes if c["numero"] == numero)


def _post_simple(client: TestClient, assoc: str, libelle: str, montant: str, jour: str):
    resp = client.post(
        f"/api/asso/{assoc}/ecritures/simple",
        json={
            "categorie_id": _categorie_id(client, assoc, libelle),
            "compte_tresorerie_id": _treso_id(client, assoc, "512"),
            "montant": montant,
            "date": jour,
        },
    )
    assert resp.status_code == 201, resp.text
    # Exports are official documents: only validated entries feed them.
    ecriture_id = resp.json()["id"]
    assert (
        client.post(
            f"/api/asso/{assoc}/ecritures/{ecriture_id}/validation"
        ).status_code
        == 200
    )


def _set_solde_initial(client: TestClient, assoc: str, montant: str, jour: str) -> None:
    resp = client.post(
        f"/api/asso/{assoc}/tresorerie/{_treso_id(client, assoc, '512')}/solde-initial",
        json={"montant": montant, "date_solde_initial": jour},
    )
    assert resp.status_code == 200, resp.text


def _create_evenement(client: TestClient, assoc: str, nom: str) -> str:
    resp = client.post(f"/api/asso/{assoc}/evenements", json={"nom": nom})
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


def _books() -> tuple[TestClient, str]:
    client, assoc = _admin_with_association("admin@example.com", "alpha")
    _post_simple(client, assoc, "Cotisations", "150.00", "2026-06-10")
    _post_simple(client, assoc, "Locations", "40.00", "2026-06-20")
    return client, assoc


# --- PDF contract -----------------------------------------------------------


def _assert_pdf(resp) -> None:
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    assert "attachment" in resp.headers["content-disposition"]
    assert resp.headers["x-content-type-options"] == "nosniff"
    assert resp.content[:4] == b"%PDF"


def test_journal_pdf():
    client, assoc = _books()
    _assert_pdf(
        client.get(
            f"/api/asso/{assoc}/exports/journal.pdf",
            params={"date_from": FROM, "date_to": TO},
        )
    )


def test_grand_livre_pdf():
    client, assoc = _books()
    _assert_pdf(
        client.get(
            f"/api/asso/{assoc}/exports/grand-livre.pdf",
            params={"date_from": FROM, "date_to": TO},
        )
    )


def test_releve_pdf():
    client, assoc = _books()
    compte_id = _treso_id(client, assoc, "512")
    resp = client.get(
        f"/api/asso/{assoc}/exports/tresorerie/{compte_id}/releve.pdf",
        params={"date_from": FROM, "date_to": TO},
    )
    _assert_pdf(resp)
    assert "releve-512" in resp.headers["content-disposition"]


def test_journal_pdf_empty_period_is_valid():
    client, assoc = _admin_with_association("empty@example.com", "beta")
    _assert_pdf(client.get(f"/api/asso/{assoc}/exports/journal.pdf"))


# --- Excel content ----------------------------------------------------------


def _load(resp) -> object:
    assert resp.status_code == 200, resp.text
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert resp.content[:2] == b"PK"  # zip signature
    return load_workbook(BytesIO(resp.content))


def test_journal_xlsx_has_headers_and_amounts():
    client, assoc = _books()
    wb = _load(
        client.get(
            f"/api/asso/{assoc}/exports/journal.xlsx",
            params={"date_from": FROM, "date_to": TO},
        )
    )
    ws = wb["Journal"]
    assert [ws.cell(row=1, column=c).value for c in range(1, 8)] == [
        "Date",
        "Pièce",
        "Journal",
        "Compte",
        "Libellé",
        "Débit",
        "Crédit",
    ]
    amounts = {cell.value for row in ws.iter_rows(min_row=2) for cell in row}
    assert 150.0 in amounts  # the cotisation débit on the bank account


def test_grand_livre_xlsx_lists_accounts():
    client, assoc = _books()
    wb = _load(
        client.get(
            f"/api/asso/{assoc}/exports/grand-livre.xlsx",
            params={"date_from": FROM, "date_to": TO},
        )
    )
    ws = wb["Grand livre"]
    comptes = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert any(label and label.startswith("512") for label in comptes)


# --- Security: scoping, treasury check, isolation ---------------------------


def test_releve_rejects_non_treasury_account():
    client, assoc = _books()
    ordinary = _compte_id(client, assoc, "756")  # produit account, not treasury
    resp = client.get(f"/api/asso/{assoc}/exports/tresorerie/{ordinary}/releve.pdf")
    assert resp.status_code == 404


def test_releve_rejects_foreign_account():
    client_a, assoc_a = _books()
    client_b, assoc_b = _admin_with_association("b@example.com", "gamma")
    compte_a = _treso_id(client_a, assoc_a, "512")
    # B asks, on its own association, for an account that belongs to A → 404.
    resp = client_b.get(f"/api/asso/{assoc_b}/exports/tresorerie/{compte_a}/releve.pdf")
    assert resp.status_code == 404


def test_exports_are_member_scoped():
    client_a, assoc_a = _books()
    client_b, _ = _admin_with_association("c@example.com", "delta")
    # B is not a member of A → 404 on every export route, no data leak.
    assert client_b.get(f"/api/asso/{assoc_a}/exports/journal.pdf").status_code == 404
    assert (
        client_b.get(f"/api/asso/{assoc_a}/exports/grand-livre.xlsx").status_code == 404
    )


# --- Compte de résultat & Bilan ANC -----------------------------------------


def test_compte_resultat_data_groups_and_totals(session: Session):
    _, assoc = _books()
    data = compte_resultat_data(session, assoc, date(2026, 6, 1), date(2026, 6, 30))
    produits = {ligne.numero: ligne.montant for ligne in data.produits}
    charges = {ligne.numero: ligne.montant for ligne in data.charges}
    assert produits["756"] == Decimal("150.00")  # Cotisations
    assert charges["613"] == Decimal("40.00")  # Locations
    assert data.resultat == Decimal("110.00")


def test_bilan_data_balances(session: Session):
    client, assoc = _admin_with_association("bilan@example.com", "bil")
    _set_solde_initial(client, assoc, "1000.00", "2026-01-01")
    _post_simple(client, assoc, "Cotisations", "150.00", "2026-06-10")
    _post_simple(client, assoc, "Locations", "40.00", "2026-06-20")

    data = bilan_data(session, assoc, date(2026, 12, 31))
    # Bank = 1000 + 150 − 40 ; report à nouveau 110 = 1000 (passif) ; result 110.
    bank = next(ligne.montant for ligne in data.actif if ligne.numero.startswith("512"))
    assert bank == Decimal("1110.00")
    assert data.resultat == Decimal("110.00")
    assert data.total_actif == data.total_passif  # the balance sheet balances


def test_compte_resultat_pdf():
    client, assoc = _books()
    _assert_pdf(
        client.get(
            f"/api/asso/{assoc}/exports/compte-resultat.pdf",
            params={"date_from": FROM, "date_to": TO},
        )
    )


def test_bilan_pdf():
    client, assoc = _books()
    _assert_pdf(
        client.get(
            f"/api/asso/{assoc}/exports/bilan.pdf", params={"date_to": "2026-12-31"}
        )
    )


# --- Bilan d'événement ------------------------------------------------------


def test_evenement_bilan_pdf():
    client, assoc = _books()
    evenement_id = _create_evenement(client, assoc, "Gala 2026")
    resp = client.post(
        f"/api/asso/{assoc}/ecritures/simple",
        json={
            "categorie_id": _categorie_id(client, assoc, "Cotisations"),
            "compte_tresorerie_id": _treso_id(client, assoc, "512"),
            "montant": "300.00",
            "date": "2026-06-15",
            "evenement_id": evenement_id,
        },
    )
    assert resp.status_code == 201, resp.text
    _assert_pdf(
        client.get(f"/api/asso/{assoc}/exports/evenements/{evenement_id}/bilan.pdf")
    )


def test_evenement_bilan_rejects_foreign_event():
    client_a, assoc_a = _books()
    client_b, assoc_b = _admin_with_association("h@example.com", "eta")
    event_a = _create_evenement(client_a, assoc_a, "Privé")
    resp = client_b.get(f"/api/asso/{assoc_b}/exports/evenements/{event_a}/bilan.pdf")
    assert resp.status_code == 404
