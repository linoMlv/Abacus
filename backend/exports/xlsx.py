"""Styled Excel workbooks (openpyxl) for the spreadsheet exports."""

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.styles.borders import Border
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="0F172A")  # ink
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")
_THIN = Side(style="thin", color="E2E8F0")
_HEADER_BORDER = Border(bottom=Side(style="thin", color="0F172A"))
_AMOUNT_FORMAT = "#,##0.00"
_DATE_FORMAT = "DD/MM/YYYY"


@dataclass
class Column:
    header: str
    kind: str = "text"  # "text" | "date" | "amount"
    width: int = 16


@dataclass
class Sheet:
    title: str
    columns: list[Column]
    rows: list[list]  # raw values (str | date | Decimal | int | None)


def _coerce(value, kind: str):
    if value is None:
        return None
    if kind == "amount":
        return float(Decimal(str(value)))
    return value


def workbook_bytes(sheets: list[Sheet]) -> bytes:
    """Build a multi-sheet workbook with bold headers, frozen panes and formats."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in sheets:
        ws = wb.create_sheet(title=sheet.title[:31])  # Excel caps tab names at 31
        for idx, col in enumerate(sheet.columns, start=1):
            cell = ws.cell(row=1, column=idx, value=col.header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _HEADER_BORDER
            ws.column_dimensions[get_column_letter(idx)].width = col.width

        for r, row in enumerate(sheet.rows, start=2):
            for c, (col, value) in enumerate(zip(sheet.columns, row), start=1):
                cell = ws.cell(row=r, column=c, value=_coerce(value, col.kind))
                cell.border = Border(bottom=_THIN)
                if col.kind == "amount":
                    cell.number_format = _AMOUNT_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                elif col.kind == "date" and isinstance(value, date):
                    cell.number_format = _DATE_FORMAT

        ws.freeze_panes = "A2"
        if sheet.rows:
            last_col = get_column_letter(len(sheet.columns))
            ws.auto_filter.ref = f"A1:{last_col}{len(sheet.rows) + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
